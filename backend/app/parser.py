"""Parse a Jira export (Excel or CSV) into a normalized list of issue dicts.

Jira's CSV/Excel exports use a fairly stable set of column headers, but casing,
spacing, and custom-field suffixes (e.g. "Custom field (Story Points)") vary
between workspaces. We do a tolerant header-match to find the columns we need.

Fallbacks
---------
- If the file has no Status column, we assume every Story / Task / Bug row is
  "Done" and every Epic / Sub-task row is excluded. A warning is emitted so
  the UI can disclose this.
- If the Assignee column is missing or entirely empty, we emit a warning so
  the user knows the per-person breakdowns will all collapse to "Unassigned".
"""
from __future__ import annotations

import csv
import io
import re
from typing import Any

from openpyxl import load_workbook

# Canonical field -> ordered list of header patterns to try (case-insensitive).
HEADER_PATTERNS: dict[str, list[str]] = {
    "issue_type": [r"^issue\s*type$", r"^type$"],
    "assignee": [r"^assignee$", r"^assignee.*name$"],
    "status": [r"^status$"],
    "story_points": [
        r"^story\s*points$",
        r"story\s*point\s*estimate",
        r"custom\s*field\s*\(story\s*points\)",
    ],
}

# Issue types we treat as "plannable work" when Status is missing.
DELIVERABLE_TYPES_WHEN_NO_STATUS = {"story", "task", "bug"}


def _normalize(s: str) -> str:
    return re.sub(r"\s+", " ", str(s)).strip().lower()


def _find_column(columns: list[str], patterns: list[str]) -> str | None:
    norm = {col: _normalize(col) for col in columns}
    for pattern in patterns:
        pat = re.compile(pattern)
        for original, n in norm.items():
            if pat.search(n):
                return original
    return None


def _coerce_story_points(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, str) and not value.strip():
        return None
    try:
        f = float(value)
    except (TypeError, ValueError):
        return None
    if f != f:  # NaN guard
        return None
    return f


def _read_rows(file_bytes: bytes, filename: str) -> tuple[list[str], list[dict[str, Any]]]:
    """Return (columns, rows) where each row is a dict keyed by column header."""
    lower = filename.lower()
    if lower.endswith(".csv") or lower.endswith(".tsv"):
        text = file_bytes.decode("utf-8-sig", errors="replace")
        sep = "\t" if lower.endswith(".tsv") else ","
        reader = csv.DictReader(io.StringIO(text), delimiter=sep)
        columns = [str(c) for c in (reader.fieldnames or [])]
        rows = [{k: v for k, v in row.items()} for row in reader]
        return columns, rows

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    it = ws.iter_rows(values_only=True)
    try:
        header_row = next(it)
    except StopIteration:
        return [], []

    columns = [str(c).strip() if c is not None else "" for c in header_row]
    rows: list[dict[str, Any]] = []
    for values in it:
        if values is None:
            continue
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in values):
            continue
        row: dict[str, Any] = {}
        for idx, col in enumerate(columns):
            if not col:
                continue
            row[col] = values[idx] if idx < len(values) else None
        rows.append(row)
    return columns, rows


def _is_present(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def parse_jira_file(file_bytes: bytes, filename: str) -> tuple[list[dict], list[str]]:
    """Return (issues, warnings).

    Each issue is a dict with keys: issue_type, assignee, status, story_points.
    """
    warnings: list[str] = []

    try:
        columns, rows = _read_rows(file_bytes, filename)
    except Exception as e:
        raise ValueError(f"Could not read file: {e}") from e

    if not rows:
        return [], ["File contains no rows."]

    mapping: dict[str, str | None] = {}
    for canonical, patterns in HEADER_PATTERNS.items():
        mapping[canonical] = _find_column(columns, patterns)

    status_missing = mapping["status"] is None
    if status_missing:
        warnings.append(
            "No Status column found. Assumed every Story/Task/Bug row is Done; "
            "Epic and Sub-task rows are excluded."
        )
    if mapping["issue_type"] is None:
        warnings.append("No Issue Type column found. All rows will be treated uniformly.")
    if mapping["assignee"] is None:
        warnings.append("No Assignee column found. Per-person breakdowns will be empty.")
    if mapping["story_points"] is None:
        warnings.append("No Story Points column found. Story-point totals will be zero.")

    issues: list[dict] = []
    for row in rows:
        issue_type = (
            str(row.get(mapping["issue_type"])).strip()
            if mapping["issue_type"] and _is_present(row.get(mapping["issue_type"]))
            else None
        )
        assignee = (
            str(row.get(mapping["assignee"])).strip()
            if mapping["assignee"] and _is_present(row.get(mapping["assignee"]))
            else None
        )
        story_points = _coerce_story_points(row.get(mapping["story_points"]) if mapping["story_points"] else None)

        if status_missing:
            # Fallback: treat plannable work as Done; anything else stays None
            # (so it won't be counted as delivered).
            if issue_type and issue_type.lower() in DELIVERABLE_TYPES_WHEN_NO_STATUS:
                status: str | None = "Done"
            else:
                status = None
        else:
            status = (
                str(row.get(mapping["status"])).strip()
                if mapping["status"] and _is_present(row.get(mapping["status"]))
                else None
            )

        issues.append(
            {
                "issue_type": issue_type,
                "assignee": assignee,
                "status": status,
                "story_points": story_points,
            }
        )

    # Extra warning if the Assignee column exists but every row is blank.
    if mapping["assignee"] is not None and all(not it["assignee"] for it in issues):
        warnings.append(
            "Assignee column is present but empty on every row. "
            "Per-person breakdowns will collapse to 'Unassigned'."
        )

    return issues, warnings
